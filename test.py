import math
import numpy as np
import pandas as pd
from pandas import DataFrame
from sklearn.cluster import KMeans
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# read data from excel and process
def read_data():
    # 由于原表格没有表头，此处设置header=None
    data = pd.read_excel(r'dataset_k.xlsx', header=None)
    # 转换为数组格式
    data = np.array(data)
    # 整理格式为一个大小为200的矩阵
    # data = data.reshape(-1)
    # data = list(data)
    return data


# write data to excel
def write_data(result):
    wb = load_workbook(r'test.xlsx')
    sheet = wb.active
    length = len(result)

    # for iu in range(length):
    #     if result[iu] == 1:
    #         result[iu] = 0
    #     else:
    #         result[iu] = 1
    sheet['A1'] = 'outputData'
    for i in range(1, length+1):
        sheet.cell(row=i+1, column=1).value = result[i-1]
    wb.save(r'test.xlsx')
    print("over!")


data = read_data()
intr = 0.51
out = []
for i in range(len(data)):
    if data[i] < intr:
        out.append(0)
    else:
        out.append(1)
print(out)

write_data(out)

