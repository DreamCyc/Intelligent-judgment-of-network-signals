import sklearn
import xlrd
import numpy as np
from sklearn.svm import SVC
from sklearn import metrics
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from openpyxl import load_workbook


# write data to excel
def write_data(result):
    wb_w = load_workbook(r'result_S.xlsx')
    sheet = wb_w.active
    length = len(result)
    sheet['A1'] = 'outputData'
    for i in range(1, length+1):
        sheet.cell(row=i+1, column=1).value = result[i-1]
    wb_w.save(r'result_S.xlsx')
    print("over!")

wb = xlrd.open_workbook(r"data_accuracy.xls")
# sheet_1 = wb.sheet_by_name('train')  # 或者
sheet_1 = wb.sheet_by_index(0)
x = np.ones((sheet_1.nrows, 1))
y = np.ones((sheet_1.nrows, 1))
for i in range(0, sheet_1.nrows):
    x[i][0] = sheet_1.cell(i, 0).value
    y[i] = sheet_1.cell(i, 1).value
x = np.array(x)
y = np.array(y)
print(x)
print(y)
print(x.shape)
print(y.shape)
print("----")

wb2 = xlrd.open_workbook(r"dataset_k.xlsx")
sheet_1 = wb2.sheet_by_index(0)
x_p = np.ones((sheet_1.nrows, 1))
for i in range(0, sheet_1.nrows):
    x_p[i][0] = sheet_1.cell(i, 0).value
x_p = np.array(x_p)
print(x_p.shape)

# 针对clf中的惩罚系数C和核函数kernel进行参数调整
clf = SVC(C=10, kernel='linear')
clf.fit(x, y)
result = clf.predict(x_p)
print(len(result))
print(result)
write_data(result)
