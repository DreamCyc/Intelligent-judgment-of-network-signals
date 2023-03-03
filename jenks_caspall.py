import random
import math
import numpy as np
import pandas as pd
# Globals variables used by both Clusters and Cluster classes
# they define how the variances will be computed
from openpyxl import load_workbook

USE_MEAN = True  # if False, use median
SQUARE_DEV = True  # if False, use absolute deviation


def getMedian(values):
    n = len(values)
    idx = n // 2
    if n % 2:  # odd number
        return values[idx]
    else:  # even number
        return (values[idx - 1] + values[idx]) / 2


class Cluster():
    def __init__(self, data, i, j):
        '''
		We define a cluster with its first and last index
		these indices refer to the original dataset
		the last index refers to a value that is include in this cluster
		'''
        self.data = data
        self.i, self.j = i, j

    def __getitem__(self, idx):
        return self.values[idx]

    def __iter__(self):
        return iter(self.values)

    def __str__(self):
        return str(self.values)

    def __len__(self):
        return len(self.values)

    @property
    def values(self):
        return self.data[self.i:self.j + 1]

    @property
    def startValue(self):
        return self.data[self.i]

    @property
    def endValue(self):
        return self.data[self.j]

    @property
    def indices(self):
        return [self.i, self.j]

    @property
    def size(self):
        return len(self.values)

    @property
    def mean(self):
        values = self.values  # for speed, avoid extracting values multiple time
        return sum(values) / len(values)

    @property
    def median(self):
        return getMedian(self.values)

    @property
    def sumSquareDev(self):
        m = self.mean  # precompute the mean for avoids slowdown
        return sum([(v - m) ** 2 for v in self.values])

    @property
    def sumAbsDev(self):
        m = self.mean
        return sum([abs(v - m) for v in self.values])

    @property
    def variance(self):
        '''
		Within classe variance
		Result depends on how globals USE_MEAN and SQUARE_DEV are defined
		It will be :
		* the sum of square deviation from class values to class mean or median
		or
		* the sum of deviation from class values to class mean or median
		'''
        if USE_MEAN:
            m = self.mean
        else:
            m = self.median
        #
        if SQUARE_DEV:
            return sum([(v - m) ** 2 for v in self.values])
        else:
            return sum([abs(v - m) for v in self.values])


class Clusters():
    # helping functions for initialization
    def buildFromBreakPtsIdx(self, breakPointsIdx):
        n, k = len(self.data), len(breakPointsIdx) + 1
        self.clusters = []
        for idx in range(k):
            if idx == 0:
                i = 0
            else:
                i = self.clusters[idx - 1].j + 1
            #
            if idx == k - 1:
                j = n - 1
            else:
                # breakPoints[i] is the last value included to ith cluster
                j = breakPointsIdx[idx]
                if j == n - 1:  # last value is in breaks list
                    j -= 1  # adjust so that the last value will be affected to last cluster
            #
            self.clusters.append(Cluster(self.data, i, j))

    def buildFromCentroidsIdx(self, centroidsIdx):
        k = len(centroidsIdx)
        breakPointsIdx = []
        for idx in range(k - 1):
            i, j = centroidsIdx[idx], centroidsIdx[idx + 1]
            m1, m2 = self.data[i], self.data[j]
            vIdx = i + 1
            while True:
                v = self.data[vIdx]
                dst1 = abs(m1 - v)
                dst2 = abs(m2 - v)
                if dst1 > dst2:
                    breakPointsIdx.append(vIdx - 1)
                    break
                else:
                    vIdx += 1
        # build clusters with these breakpoints
        self.buildFromBreakPtsIdx(breakPointsIdx)

    def values2Idx(self, values, findLastIdx=False):
        if not findLastIdx:
            return [self.data.index(v) for v in values]
        else:
            rvData = list(reversed(self.data))
            return [rvData.index(v) for v in values]

    def __init__(self, data, k, style='quantile'):
        '''
		Create k clusters with an initial classification
		'''

        self.data = data
        self.data.sort()
        n = len(self.data)

        # precompute data statistics
        self.dataMean = sum(self.data) / n
        self.dataMedian = getMedian(self.data)
        self.dataSumSquareDev = sum([(self.dataMean - v) ** 2 for v in self.data])
        self.dataSumAbsDev = sum([abs(self.dataMean - v) for v in self.data])
        self.dataSumSquareDevMedian = sum([(self.dataMedian - v) ** 2 for v in self.data])
        self.dataSumAbsDevMedian = sum([abs(self.dataMedian - v) for v in self.data])

        # a little hack to build clusters from breaks values
        # use it only for testing a predefined partition
        if type(k) is list:
            breakPoints = k
            breakPointsIdx = self.values2Idx(breakPoints)
            self.buildFromBreakPtsIdx(breakPointsIdx)
            return

        if not 0 < k < n:
            raise ValueError('Wrong expected number of classes')
        if style not in ['quantile', 'equal_interval', 'random', 'kpp', 'max']:
            print('Incorrect requested init style, use default style')
            style = 'quantile'

        # request only one classe
        if k == 1:
            self.clusters = [Cluster(self.data, 0, n - 1)]

        elif style == 'quantile':
            # quantile = number of value per clusters
            q = int(n // k)  # floor division
            if q == 1:
                raise ValueError('Too many expected classes')
            # Make a list of Cluster object
            self.clusters = [Cluster(self.data, i, i + q - 1) for i in range(0, q * k, q)]
            #  adjust the last index of the last cluster to the effective number of value
            self.clusters[-1].j = n - 1

        elif style == 'equal_interval':
            mini, maxi = self.data[0], self.data[-1]
            delta = maxi - mini
            interval = delta / k
            breakPointsIdx = []
            target = mini + interval
            for i, v in enumerate(self.data):
                if len(breakPointsIdx) == k - 1:
                    break
                if v > target:
                    breakPointsIdx.append(i - 1)
                    target += interval
            # build clusters with these breakpoints
            self.buildFromBreakPtsIdx(breakPointsIdx)

        elif style == 'random':
            # generate random indices
            breakPointsIdx = random.sample(range(0, n - 1), k)
            breakPointsIdx.sort()
            # build clusters with them as breakpoints
            self.buildFromBreakPtsIdx(breakPointsIdx)

        elif style == 'kpp':
            # # kmeans++ initialization
            # # this code is based on an example describe at
            # # http://stackoverflow.com/questions/5466323/how-exactly-does-k-means-work
            #
            # use kpp to init centroids or directly breaksPoints ?
            AS_CENTROIDS = True
            if AS_CENTROIDS:
                # to get k classes we need k centroids
                n_init = k
            else:
                # to get k classes we need k-1 breakpoints
                n_init = k - 1
            #
            # pick up a random value as first break value
            centroidsIdx = random.sample(range(n), 1)
            # n_init - 1 values remaining to find
            for cpt in range(n_init - 1):
                centroidsValues = [self.data[idx] for idx in centroidsIdx]
                # For each value compute the square distance to the nearest centroid
                dst = [min([(c - v) ** 2 for c in centroidsValues]) for v in self.data]
                # compute probability of each values
                sumDst = sum(dst)
                probs = [d / sumDst for d in dst]
                # compute the cumulative probability (range from 0 to 1)
                cumSum = 0  # cumulative sum
                cumProbs = []  # cumulative proba
                for p in probs:
                    cumSum += p
                    cumProbs.append(cumSum)
                # now try to find a new centroid
                find = False
                while not find:
                    # generate a random probability (a float range from 0 to 1)
                    r = random.random()
                    # loop over the probability of each value...
                    for i, p in enumerate(cumProbs):
                        # ...and search for the first value with a probability higher than the random one
                        if r < p:
                            # add idx to our centroids list if it's not already there
                            if i not in centroidsIdx:
                                centroidsIdx.append(i)
                                find = True
                                break
            centroidsIdx.sort()
            # find the breakpoints corresponding to these centroids
            if AS_CENTROIDS:
                self.buildFromCentroidsIdx(centroidsIdx)
            else:
                # build our clusters with these centroids as breakpoints
                self.buildFromBreakPtsIdx(centroidsIdx)
        # print(centroidsIdx)
        # print(self.indices)

        elif style == 'max':
            # a simple method to get well spaced splits
            # because data is ordered, start with first and last values
            breakPointsIdx = [0, n - 1]
            for cpt in range(k - 1):
                breaksValues = [self.data[idx] for idx in breakPointsIdx]
                # For each value compute the square distance to the nearest centroid
                dst = [min([(c - v) ** 2 for c in breaksValues]) for v in self.data]
                # choose the value that has the greatest minimum-distance to the previously selected centers
                idx = dst.index(max(dst))
                breakPointsIdx.append(idx)
            # Exclude first and last values
            breakPointsIdx.sort()
            breakPointsIdx = breakPointsIdx[1:-1]
            # build our clusters with these centroids as breakpoints
            self.buildFromBreakPtsIdx(breakPointsIdx)

    def __str__(self):
        return str([c.values for c in self.clusters])

    def __getitem__(self, idx):
        return self.clusters[idx]

    def __setitem__(self, idx, cluster):
        if isinstance(cluster, Cluster):
            self.clusters[idx] = cluster
        else:
            raise ValueError('Requiere a Cluster instance not %s' % type(cluster))
        if not self.checkIntegrity():
            raise ValueError('Incorrect cluster definition for this position')

    def __iter__(self):
        return iter(self.clusters)

    def __len__(self):
        return len(self.clusters)

    @property
    def k(self):
        return len(self.clusters)

    @property
    def size(self):
        return len(self.clusters)

    @property
    def n(self):
        return len(self.data)

    @property
    def values(self):
        return [c.values for c in self.clusters]

    @property
    def indices(self):
        return [c.indices for c in self.clusters]

    @property
    def breaks(self):
        return [self.data[c.j] for c in self.clusters[:-1]]

    @property
    def breaksWithBounds(self):
        return [self.data[0]] + [self.data[c.j] for c in self.clusters]

    @property
    def dataVariance(self):
        if SQUARE_DEV:
            if USE_MEAN:
                return self.dataSumSquareDev
            else:
                return self.dataSumSquareDevMedian
        else:
            if USE_MEAN:
                return self.dataSumAbsDev
            else:
                return self.dataSumAbsDevMedian

    @property
    def withinVariances(self):
        return [c.variance for c in self.clusters]

    @property
    def sumWithinVariances(self):
        '''Sum of withim clusters sum square dev or sum abs dev'''
        return sum([c.variance for c in self.clusters])

    @property
    def betweenVariance(self):
        return self.dataVariance - self.sumWithinVariances

    @property
    def betweenVariance_calc(self):
        '''
		Between classes variance
		Sum of square deviation from classes means to total mean
		'''
        if SQUARE_DEV:
            if USE_MEAN:
                return sum([c.size * (self.dataMean - c.mean) ** 2 for c in self.clusters])
            else:
                return sum([c.size * (self.dataMedian - c.median) ** 2 for c in self.clusters])
        else:
            if USE_MEAN:
                return sum([c.size * abs(self.dataMean - c.mean) for c in self.clusters])
            else:
                return sum([c.size * abs(self.dataMedian - c.median) for c in self.clusters])

    @property
    def gvf(self):
        '''
		Goodness of Variance Fit
		ranges from 1 (perfect fit) to 0 (awful fit)
		if not SQUARE_DEV then this will be the Tabular Accuracy Index
		'''
        return 1 - (self.sumWithinVariances / self.dataVariance)

    def printStats(self):
        print('%i values, %i classes' % (self.n, self.k))
        # print("Breaks %s" %self.breaks)
        print("Breaks with bounds %s" % self.breaksWithBounds)
        # print("Data Variance %i" %self.dataVariance)
        print("Sum of within classes variances %i" % self.sumWithinVariances)
        print("Between classes variance %i" % self.betweenVariance)
        print("Goodness of variance fit %f" % self.gvf)

    def checkIntegrity(self):
        # last index +1 of each cluster == start index of the next cluster
        return all([c.j + 1 == self.clusters[idx + 1].i for idx, c in enumerate(self.clusters[:-1])])

    def moveForward(self, idx):
        '''
		Move last value of a given cluster index to its next cluster
		'''
        if idx == len(self.clusters) - 1:
            # last cluster, cannot move forward
            return False
        # if self.clusters[idx].i == self.clusters[idx].j:
        if self.clusters[idx].size == 1:
            # only one value remaining in this cluster
            # do not execute this move to avoid having an empty cluster
            return False
        # decrease right border index of current cluster
        self.clusters[idx].j -= 1
        # decrease left border index of the next cluster
        self.clusters[idx + 1].i -= 1
        return True

    def moveBackward(self, idx):
        '''
		Move first value of a given cluster index to its previous cluster
		'''
        if idx == 0:
            # first cluster, cannot move backward
            return False
        if self.clusters[idx].size == 1:
            # only one value remaining in this cluster
            # do not execute this move to avoid having an empty cluster
            return False
        # increase left border index of the current cluster
        self.clusters[idx].i += 1
        # increase right border index of previous cluster
        self.clusters[idx - 1].j += 1
        return True

    def save(self):
        self.previousIndices = self.indices

    def restore(self):
        for idx, c in enumerate(self.clusters):
            c.i, c.j = self.previousIndices[idx]


# -------------------


def kmeans1D(clusters, updateMeanAfterEachMove=False):
    nbIter = 0
    nbMoves = 0
    changeOccured = True
    while True:
        nbIter += 1
        nbMovesIter = 0
        changeOccured = False

        # save actual partition and sum of within variances
        sumWithinVariances = clusters.sumWithinVariances
        clusters.save()

        # if not updateMeanAfterEachMove, keep the same initial means during this iteration
        # means are re-computed only after all the data points have been assigned to their nearest centroids
        means = [c.mean for c in clusters]

        # for each border...
        k = len(clusters)
        for idx in range(k - 1):
            c1, c2 = clusters[idx], clusters[idx + 1]
            # m1, m2 = means[idx], means[idx+1]
            adjusted = False

            # try to adjust this border by moving forward (c1 -> c2)
            while True:
                breakValue = c1.endValue
                # get distance to means
                dst1 = abs(breakValue - means[idx])
                dst2 = abs(breakValue - means[idx + 1])
                if dst1 > dst2:  # this value will be better in c2
                    if clusters.moveForward(idx):  # move is a success ...
                        adjusted = True
                        nbMovesIter += 1
                        if updateMeanAfterEachMove:  # always use an updated mean
                            means[idx], means[idx + 1] = c1.mean, c2.mean
                    else:
                        break
                else:
                    break

            if not adjusted:
                # maybee we can do it backward (c1 <- c2)
                while True:
                    breakValue = c2.startValue
                    dst1 = abs(breakValue - means[idx])
                    dst2 = abs(breakValue - means[idx + 1])
                    if dst2 > dst1:
                        if clusters.moveBackward(idx + 1):
                            adjusted = True
                            nbMovesIter += 1
                            if updateMeanAfterEachMove:  # always use an updated mean
                                means[idx], means[idx + 1] = c1.mean, c2.mean
                        else:
                            break
                    else:
                        break

            if adjusted:
                changeOccured = True

        if not changeOccured:
            break
        elif clusters.sumWithinVariances > sumWithinVariances:
            # This new partition isn't better, so restaure the previous one and break the loop
            clusters.restore()
            break
        else:
            nbMoves += nbMovesIter

    return nbIter, nbMoves


# -------------------


def forceCycle(clusters):
    k = len(clusters)
    nbMoves = 0

    # store a list of variances (because we don't want to recompute all variance after each moves)
    # instead of GVF, we'll use variance (sum of squared or absolute deviation) as evaluation criterion
    # so we can just update and compare the sum of dev for the modified clusters
    # it will be faster than use GVF, and the results remaining the same
    sumdev = clusters.withinVariances

    # backward forcing cycle.
    # from low to high classes, move backward a value
    for idx in range(1, k):
        c1, c2 = clusters[idx - 1], clusters[idx]
        while True:
            #
            previousSumDev = sumdev[idx - 1] + sumdev[idx]
            # forcing first value of a given cluster index to its previous cluster
            moved = clusters.moveBackward(idx)
            if not moved:
                break
            newSumDev = [c1.variance, c2.variance]
            if sum(newSumDev) > previousSumDev:
                # undo last move and break loop
                clusters.moveForward(idx - 1)
                break
            else:
                sumdev[idx - 1], sumdev[idx] = newSumDev
                nbMoves += 1

    # forward forcing cycle
    # from high to low classes, move forward a value
    for idx in range(k - 2, -1, -1):
        c1, c2 = clusters[idx], clusters[idx + 1]
        while True:
            previousSumDev = sumdev[idx] + sumdev[idx + 1]
            # forcing last value of a given cluster index to its next cluster
            moved = clusters.moveForward(idx)
            if not moved:
                break
            newSumDev = [c1.variance, c2.variance]
            if sum(newSumDev) > previousSumDev:
                # undo last move and break loop
                clusters.moveBackward(idx + 1)
                break
            else:
                sumdev[idx], sumdev[idx + 1] = newSumDev
                nbMoves += 1

    return nbMoves


# -------------------


def jenksCaspall(data, k, nbAttempt=1, initStyle='kmeanpp'):
    bestClusters = None
    for i in range(nbAttempt):

        print('Running Jenks-Caspall natural breaks...')
        print('**Attempt number %i' % (i + 1))

        # kmean++ intit
        clusters = Clusters(data, k, initStyle)
        print('Step 1 : kmeans++ initalization, GVF = %f' % clusters.gvf)
        # print(clusters.breaks)

        # kmeans
        nbIter, nbMoves = kmeans1D(clusters)
        print('Step 2 : kmeans complete in %i iterations and %i moves, GVF = %f' % (nbIter, nbMoves, clusters.gvf))
        # print(clusters.breaks)
        # force cycle
        nbForceCycle = 0
        nbMovesAll = 0
        while True:
            nbForceCycle += 1
            nbMoves = forceCycle(clusters)
            if not nbMoves:
                break
            else:
                nbMovesAll += nbMoves
        print(
            'Step 3 : Forcing completed in %i cycles and %i moves, GVF = %f' % (nbForceCycle, nbMovesAll, clusters.gvf))
        # print(clusters.breaks)
        # Assign best partition
        if i == 0:
            bestClusters = clusters
        else:
            if clusters.gvf > bestClusters.gvf:
                bestClusters = clusters
    # Finish
    print('Jenks-Caspall competed!')
    bestClusters.printStats()
    return bestClusters


# 读取数据
# read data from excel and process
def read_data():
    # 由于原表格没有表头，此处设置header=None
    data = pd.read_excel(r'dataset_k.xlsx', header=None)
    data = np.array(data)
    # 整理格式为一个大小为200的矩阵
    # data = data.reshape(-1)
    # data = list(data)
    return data


# get 输出结果
def class_get(input, inter):
    out = []
    for item in input:
        if item < inter:
            out.append(0)
        else:
            out.append(1)
    return out


# 结果写入Excel-result_J
# write data to excel
def write_data(result):
    wb_w = load_workbook(r'result_J.xlsx')
    sheet = wb_w.active
    length = len(result)
    sheet['A1'] = 'outputData'
	# 在第一行的信息输入完毕后，逐行写入数据
    for i in range(1, length + 1):
        sheet.cell(row=i + 1, column=1).value = result[i - 1]
    wb_w.save(r'result_J.xlsx')
    print("over!")


if __name__ == '__main__':
    # This is the original dataset used by Jenks and Caspall in their paper
    # Jenks reach a TAI of 0.73416 with the following breaks [41.2, 58.5, 75.51, 100.10]
    # 获取数据并处理
    data = read_data()
    print(data)
    data = data.reshape(-1)
    # 分类数2，步数为4
    k = 2
    nbAttempt = 4
    # clusters = jenksCaspall(data, k, nbAttempt, initStyle='kpp')
    clusters = jenksCaspall(data, k, nbAttempt, initStyle='equal_interval')

    print(clusters.breaks)
    # 取类别中间的值作为均衡
    inter = clusters.breaks
    # 得到判决结果
    out_result = class_get(data, inter)
    print(out_result)
    # 写入Excel
    write_data(out_result)