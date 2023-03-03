import math
import numpy as np
import pandas as pd
from pandas import DataFrame
from sklearn.cluster import KMeans
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# read data from excel and process
def read_data():
    # 由于原表格没有表头，此处设置header=None
    data = pd.read_excel(r'dataset_k.xlsx', header=None)
    # 转换为数组格式
    data = np.array(data)
    # 整理格式为一个大小为200的矩阵
    # data = data.reshape(-1)
    # data = list(data)
    return data


# write data to excel
def write_data(result):
    wb = load_workbook(r'result_k.xlsx')
    sheet = wb.active
    length = len(result)

    # for iu in range(length):
    #     if result[iu] == 1:
    #         result[iu] = 0
    #     else:
    #         result[iu] = 1
    sheet['A1'] = 'outputData'
    for i in range(1, length+1):
        sheet.cell(row=i+1, column=1).value = result[i-1]
    wb.save(r'result_k.xlsx')
    print("over!")


# 一种非线性映射的测试
def softmax(x):
    s = []
    x_sum = 0
    for i in range(len(x)):
        x_sum = x_sum + np.exp(x[i])
    for i in range(len(x)):
        x_exp = np.exp((-1)*x[i])
        s.append(x_exp / x_sum)
    return s


# 一种非线性映射方法的测试
def nonline(x):
    for i in range(len(x)):
        if x[i] < 0:
            x[i] = pow(2, (-1)*x[i]) + 1 - 1/math.sqrt(2)
            x[i] = (-1)*x[i]
        else:
            x[i] = math.log(x[i], 2)
    return x


if __name__ == '__main__':
    # k-means algorithm
    # get dataset
    data_k = read_data()
    max_num = max(data_k)
    print(max_num)
    min_num = min(data_k)
    print(min_num)
    num = []

    # 以下注释内容为过程中测试的线性函数和非线性函数
    # data_k = 4*(data_k-0.5)
    # data_k = list(data_k)
    # data_k = softmax(data_k)
    # data_k = np.array(data_k)
    # print(type(data_k))
    # for i in range(200):
    #     num.append(-1+i*0.02)
    # plt.scatter(data_k, data_k)
    # plt.show()
    # print(type(data_k[0]))
    # print(len(data_k[0]))
    # data_k = np.exp(data_k)-1

    # for i in data_k:
    #     if i < 0.5:
    #         data_k = pow(2, (-1)*data_k) + 1 - 1/math.sqrt(2)
    #     else:
    #         data_k = (-1) * math.log(data_k, 2)
    # data_k = 2 * (data_k+1)
    # data_k = softmax(data_k)
    # data_k = np.math.log2(data_k)
    # data_k = np.exp((-1)*data_k)
    # data_k = np.sqrt(9-pow(data_k, 2))
    estimator = KMeans(n_clusters=2)
    estimator.fit(data_k)
    label = estimator.labels_
    center = estimator.cluster_centers_
    inertia = estimator.inertia_
    print(label)
    print(center)
    print(inertia)
    write_data(label)