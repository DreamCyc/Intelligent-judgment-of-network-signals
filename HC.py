import xlrd
from sklearn.cluster import AgglomerativeClustering
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import scipy.cluster.hierarchy as shc

# 数据读取
wb = xlrd.open_workbook(r"dataset_k.xlsx")
sheet_1 = wb.sheet_by_index(0)
x = np.ones((sheet_1.nrows, 1))
for i in range(0, sheet_1.nrows):
    x[i][0] = sheet_1.cell(i, 0).value
x = np.array(x)
print(x)
print(x.shape)


# read data from excel and process
def read_data():
    # 由于原表格没有表头，此处设置header=None
    data = pd.read_excel(r'dataset_k.xlsx', header=None)
    # 转换为数组格式
    data = np.array(data)
    # 整理格式为一个大小为200的矩阵
    # data = data.reshape(-1)
    # data = list(data)
    return data


# write data to excel
def write_data(result):
    wb_w = load_workbook(r'result_H.xlsx')
    sheet = wb_w.active
    length = len(result)

    # for iu in range(length):
    #     if result[iu] == 1:
    #         result[iu] = 0
    #     else:
    #         result[iu] = 1
    sheet['A1'] = 'outputData'
    for i in range(1, length+1):
        sheet.cell(row=i+1, column=1).value = result[i-1]
    wb_w.save(r'result_H.xlsx')
    print("over!")

# ['ward', 'complete', 'average', 'single']
# 拟合并写入数据
clustering = AgglomerativeClustering(n_clusters=2, linkage='ward').fit(x)
print(clustering.labels_)
write_data(clustering.labels_)

# 绘制层次聚类图
plt.figure(figsize=(10, 7))
plt.title("Customer Presentation")
dend = shc.dendrogram(shc.linkage(x, method='ward'))
plt.savefig("./Hc-result.jpg")
plt.show()
